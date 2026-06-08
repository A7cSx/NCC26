from fastapi import FastAPI, APIRouter, HTTPException, Header, Depends, Request
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import re
import logging
import bcrypt
import jwt
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'ncc-admin-2026')
JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALG = "HS256"
JWT_TTL_DAYS = 7

app = FastAPI()
api_router = APIRouter(prefix="/api")


# ---------- Models ----------
class Employee(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    username: str
    full_name: str
    name: str  # kept for back-compat (= full_name)
    password_hash: str
    email: Optional[str] = None
    phone: Optional[str] = None
    notify_enabled: bool = True
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class RegisterPayload(BaseModel):
    full_name: str
    username: str
    employee_id: str
    password: str
    confirm_password: str


class LoginPayload(BaseModel):
    username: str
    password: str


class EmployeeUpdate(BaseModel):
    email: Optional[str] = None
    phone: Optional[str] = None
    notify_enabled: Optional[bool] = None


class Match(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    team_a: str
    team_b: str
    team_a_ar: str
    team_b_ar: str
    flag_a: str = ""
    flag_b: str = ""
    group: str = ""
    stage: str = "Group Stage"
    kickoff: str  # ISO datetime
    status: str = "upcoming"  # upcoming | live | finished
    result_a: Optional[int] = None
    result_b: Optional[int] = None
    winner: Optional[str] = None  # team_a | team_b | draw
    stream_url: Optional[str] = None  # YouTube embed URL or any iframe-able stream
    venue: str = ""


class MatchCreate(BaseModel):
    team_a: str
    team_b: str
    team_a_ar: str = ""
    team_b_ar: str = ""
    flag_a: str = ""
    flag_b: str = ""
    group: str = ""
    stage: str = "Group Stage"
    kickoff: str


class MatchResult(BaseModel):
    result_a: int
    result_b: int


class Prediction(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    match_id: str
    winner: str  # team_a | team_b | draw
    score_a: int
    score_b: int
    points: int = 0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class PredictionCreate(BaseModel):
    employee_id: str
    match_id: str
    winner: str
    score_a: int
    score_b: int


# Notification = a scheduled or fired reminder. Backend GET/POST endpoints expose this store.
class Notification(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    match_id: str
    title: str
    body: str = ""
    fire_at: str  # ISO datetime when the reminder should be shown
    channel: str = "in_app"  # in_app | email | whatsapp | telegram
    target: str = "all"  # "all" or specific employee_id
    sent: bool = False
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class NotificationCreate(BaseModel):
    match_id: str
    title: str
    body: str = ""
    fire_at: str
    channel: str = "in_app"
    target: str = "all"


# ---------- Helpers ----------
def clean(doc):
    if doc is None:
        return None
    doc.pop('_id', None)
    doc.pop('password_hash', None)
    return doc


def verify_admin(x_admin_password: str = Header(None)):
    if x_admin_password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Unauthorized")
    return True


def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def create_token(employee_id: str, username: str) -> str:
    payload = {
        "sub": employee_id,
        "username": username,
        "exp": datetime.now(timezone.utc) + timedelta(days=JWT_TTL_DAYS),
        "iat": datetime.now(timezone.utc),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


async def get_current_user(request: Request) -> dict:
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(401, "Not authenticated")
    token = auth[7:]
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")
    emp = await db.employees.find_one({"employee_id": payload["sub"]}, {"_id": 0, "password_hash": 0})
    if not emp:
        raise HTTPException(401, "User not found")
    return emp


USERNAME_RE = re.compile(r"^[a-zA-Z0-9_.-]{3,32}$")


def compute_points(p_winner: str, p_a: int, p_b: int, r_winner: str, r_a: int, r_b: int) -> int:
    if p_a == r_a and p_b == r_b:
        return 5  # exact score
    if p_winner == r_winner:
        return 3  # correct winner
    return 0


def determine_winner(a: int, b: int) -> str:
    if a > b:
        return "team_a"
    if b > a:
        return "team_b"
    return "draw"


LOCK_MINUTES_BEFORE_KICKOFF = 5


def is_locked(match: dict) -> bool:
    """A match is locked for predictions if it's within LOCK_MINUTES_BEFORE_KICKOFF
    of kickoff, or if the status is no longer 'upcoming'."""
    if match.get("status") != "upcoming":
        return True
    try:
        ko = datetime.fromisoformat(match["kickoff"].replace("Z", "+00:00"))
    except Exception:
        return False
    lock_time = ko - timedelta(minutes=LOCK_MINUTES_BEFORE_KICKOFF)
    return datetime.now(timezone.utc) >= lock_time


# ---------- Auth / Employees ----------
@api_router.post("/auth/register")
async def register(payload: RegisterPayload):
    full_name = payload.full_name.strip()
    username = payload.username.strip().lower()
    eid = payload.employee_id.strip()
    pw = payload.password
    confirm = payload.confirm_password

    if not full_name or len(full_name) < 3:
        raise HTTPException(400, "FULL_NAME_REQUIRED")
    if not USERNAME_RE.match(username):
        raise HTTPException(400, "USERNAME_INVALID")
    if not eid:
        raise HTTPException(400, "EMPLOYEE_ID_REQUIRED")
    if not pw or len(pw) < 6:
        raise HTTPException(400, "PASSWORD_TOO_SHORT")
    if pw != confirm:
        raise HTTPException(400, "PASSWORD_MISMATCH")

    if await db.employees.find_one({"username": username}):
        raise HTTPException(409, "USERNAME_TAKEN")
    if await db.employees.find_one({"employee_id": eid}):
        raise HTTPException(409, "EMPLOYEE_ID_TAKEN")

    emp = Employee(
        employee_id=eid,
        username=username,
        full_name=full_name,
        name=full_name,
        password_hash=hash_password(pw),
    )
    await db.employees.insert_one(emp.model_dump())
    token = create_token(eid, username)
    user_out = emp.model_dump()
    user_out.pop("password_hash", None)
    return {"token": token, "user": user_out}


@api_router.post("/auth/login")
async def login(payload: LoginPayload):
    username = payload.username.strip().lower()
    emp = await db.employees.find_one({"username": username})
    if not emp:
        raise HTTPException(401, "INVALID_CREDENTIALS")
    if not verify_password(payload.password, emp.get("password_hash", "")):
        raise HTTPException(401, "INVALID_CREDENTIALS")
    token = create_token(emp["employee_id"], emp["username"])
    return {"token": token, "user": clean(emp)}


@api_router.get("/auth/me")
async def auth_me(current=Depends(get_current_user)):
    return current


@api_router.patch("/employees/{employee_id}")
async def update_employee(employee_id: str, payload: EmployeeUpdate):
    """Update notification preferences and contact info."""
    upd = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not upd:
        raise HTTPException(400, "Nothing to update")
    res = await db.employees.update_one({"employee_id": employee_id}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(404, "Employee not found")
    emp = await db.employees.find_one({"employee_id": employee_id}, {"_id": 0})
    return emp


@api_router.get("/employees")
async def list_employees():
    emps = await db.employees.find({}, {"_id": 0}).to_list(10000)
    return emps


@api_router.get("/employees/{employee_id}")
async def get_employee(employee_id: str):
    emp = await db.employees.find_one({"employee_id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(404, "Employee not found")
    return emp


@api_router.get("/records/employee/{employee_id}")
async def employee_records(employee_id: str):
    """Comprehensive snapshot of an employee's records: profile, predictions count,
    total points (finished matches only), exact/correct counts."""
    emp = await db.employees.find_one({"employee_id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(404, "Employee not found")
    preds = await db.predictions.find({"employee_id": employee_id}, {"_id": 0}).to_list(10000)
    finished_ids = {m["id"] for m in await db.matches.find({"status": "finished"}, {"_id": 0, "id": 1}).to_list(2000)}
    pts = sum(p.get("points", 0) for p in preds if p["match_id"] in finished_ids)
    exact = sum(1 for p in preds if p.get("points", 0) == 5)
    correct = sum(1 for p in preds if p.get("points", 0) == 3)
    return {
        "employee": emp,
        "total_predictions": len(preds),
        "points": pts,
        "exact_scores": exact,
        "correct_winners": correct,
        "predictions": preds,
    }


# ---------- Matches ----------
@api_router.get("/matches")
async def list_matches():
    matches = await db.matches.find({}, {"_id": 0}).sort("kickoff", 1).to_list(1000)
    return matches


@api_router.get("/matches/{match_id}")
async def get_match(match_id: str):
    m = await db.matches.find_one({"id": match_id}, {"_id": 0})
    if not m:
        raise HTTPException(404, "Not found")
    return m


@api_router.post("/admin/matches", dependencies=[Depends(verify_admin)])
async def create_match(payload: MatchCreate):
    m = Match(**payload.model_dump())
    await db.matches.insert_one(m.model_dump())
    return m.model_dump()


@api_router.delete("/admin/matches/{match_id}", dependencies=[Depends(verify_admin)])
async def delete_match(match_id: str):
    await db.matches.delete_one({"id": match_id})
    await db.predictions.delete_many({"match_id": match_id})
    return {"ok": True}


@api_router.post("/admin/matches/{match_id}/status", dependencies=[Depends(verify_admin)])
async def set_status(match_id: str, status: str):
    if status not in ("upcoming", "live", "finished"):
        raise HTTPException(400, "invalid status")
    await db.matches.update_one({"id": match_id}, {"$set": {"status": status}})
    return {"ok": True}


@api_router.post("/admin/matches/{match_id}/result", dependencies=[Depends(verify_admin)])
async def submit_result(match_id: str, result: MatchResult):
    match = await db.matches.find_one({"id": match_id})
    if not match:
        raise HTTPException(404, "Match not found")
    r_winner = determine_winner(result.result_a, result.result_b)
    await db.matches.update_one(
        {"id": match_id},
        {"$set": {
            "result_a": result.result_a,
            "result_b": result.result_b,
            "winner": r_winner,
            "status": "finished",
            "finished_at": datetime.now(timezone.utc).isoformat(),
        }},
    )
    # Recompute points for all predictions on this match
    preds = await db.predictions.find({"match_id": match_id}).to_list(10000)
    for p in preds:
        pts = compute_points(p["winner"], p["score_a"], p["score_b"], r_winner, result.result_a, result.result_b)
        await db.predictions.update_one({"id": p["id"]}, {"$set": {"points": pts}})
    return {"ok": True, "predictions_updated": len(preds)}


@api_router.get("/admin/matches/{match_id}/predictions", dependencies=[Depends(verify_admin)])
async def admin_match_predictions(match_id: str):
    """Admin view: all predictions for a match with employee name, ID, and timestamp."""
    preds = await db.predictions.find({"match_id": match_id}, {"_id": 0}).sort("created_at", 1).to_list(10000)
    emp_ids = list({p["employee_id"] for p in preds})
    emps = await db.employees.find({"employee_id": {"$in": emp_ids}}, {"_id": 0}).to_list(10000)
    emp_map = {e["employee_id"]: e["name"] for e in emps}
    for p in preds:
        p["employee_name"] = emp_map.get(p["employee_id"], p["employee_id"])
    return {"predictions": preds, "count": len(preds)}


@api_router.get("/winners/latest")
async def latest_winners():
    """Returns the most recently finished match (by finished_at, fallback to kickoff) + all employees who scored points on it."""
    finished = await db.matches.find({"status": "finished"}, {"_id": 0}).to_list(2000)
    if not finished:
        return {"match": None, "winners": []}
    # Sort by finished_at desc (if present), then kickoff desc
    finished.sort(key=lambda m: (m.get("finished_at") or "", m.get("kickoff") or ""), reverse=True)
    match = finished[0]
    preds = await db.predictions.find(
        {"match_id": match["id"], "points": {"$gt": 0}}, {"_id": 0}
    ).sort("points", -1).to_list(10000)
    emp_ids = list({p["employee_id"] for p in preds})
    emps = await db.employees.find({"employee_id": {"$in": emp_ids}}, {"_id": 0}).to_list(10000)
    emp_map = {e["employee_id"]: e["name"] for e in emps}
    winners = []
    for p in preds:
        winners.append({
            "employee_id": p["employee_id"],
            "name": emp_map.get(p["employee_id"], p["employee_id"]),
            "score_a": p["score_a"],
            "score_b": p["score_b"],
            "points": p["points"],
            "exact": p["points"] == 5,
        })
    return {"match": match, "winners": winners}


# ---------- Stats ----------
@api_router.get("/stats/teams")
async def team_stats():
    """Aggregate prediction stats per team across all matches:
    how many employees predicted that team to win at least one match,
    and total winning votes received across all matches.
    """
    preds = await db.predictions.find({}, {"_id": 0}).to_list(200000)
    matches = await db.matches.find({}, {"_id": 0}).to_list(2000)
    match_map = {m["id"]: m for m in matches}

    # Aggregate: team_name -> {votes, employees_set, matches_set}
    team_data = {}
    for p in preds:
        m = match_map.get(p["match_id"])
        if not m:
            continue
        if p["winner"] == "team_a":
            t_en = m["team_a"]
            t_ar = m.get("team_a_ar", t_en)
            flag = m.get("flag_a", "")
        elif p["winner"] == "team_b":
            t_en = m["team_b"]
            t_ar = m.get("team_b_ar", t_en)
            flag = m.get("flag_b", "")
        else:
            continue
        td = team_data.setdefault(t_en, {
            "team": t_en, "team_ar": t_ar, "flag": flag,
            "votes": 0, "voters": set(), "matches": set(),
        })
        td["votes"] += 1
        td["voters"].add(p["employee_id"])
        td["matches"].add(p["match_id"])

    items = []
    for td in team_data.values():
        items.append({
            "team": td["team"],
            "team_ar": td["team_ar"],
            "flag": td["flag"],
            "votes": td["votes"],
            "unique_voters": len(td["voters"]),
            "matches_predicted_in": len(td["matches"]),
        })
    items.sort(key=lambda x: (-x["votes"], -x["unique_voters"], x["team"]))
    return {"teams": items, "total_predictions": len(preds)}


@api_router.get("/stats/match/{match_id}")
async def match_stats(match_id: str):
    """Predictions breakdown for a match: votes for team_a / draw / team_b,
    plus most-predicted score. Returns even for upcoming matches (shows 'crowd pick').
    """
    match = await db.matches.find_one({"id": match_id}, {"_id": 0})
    if not match:
        raise HTTPException(404, "Match not found")
    preds = await db.predictions.find({"match_id": match_id}, {"_id": 0}).to_list(20000)
    votes = {"team_a": 0, "draw": 0, "team_b": 0}
    score_counts = {}
    for p in preds:
        votes[p["winner"]] = votes.get(p["winner"], 0) + 1
        key = f"{p['score_a']}-{p['score_b']}"
        score_counts[key] = score_counts.get(key, 0) + 1
    total = len(preds)
    popular_score = max(score_counts.items(), key=lambda x: x[1])[0] if score_counts else None
    return {
        "match_id": match_id,
        "team_a": match["team_a"], "team_b": match["team_b"],
        "total": total,
        "votes": votes,
        "pct": {
            "team_a": round(votes["team_a"] * 100 / total, 1) if total else 0,
            "draw": round(votes["draw"] * 100 / total, 1) if total else 0,
            "team_b": round(votes["team_b"] * 100 / total, 1) if total else 0,
        },
        "popular_score": popular_score,
        "popular_score_count": score_counts.get(popular_score, 0) if popular_score else 0,
    }


@api_router.get("/stats/overview")
async def overview_stats():
    """Top-level numbers for marketing strip / admin dashboard."""
    employees = await db.employees.count_documents({})
    predictions = await db.predictions.count_documents({})
    matches_total = await db.matches.count_documents({})
    finished = await db.matches.count_documents({"status": "finished"})
    upcoming = await db.matches.count_documents({"status": "upcoming"})
    return {
        "employees": employees,
        "predictions": predictions,
        "matches_total": matches_total,
        "matches_finished": finished,
        "matches_upcoming": upcoming,
    }


# ---------- Notifications (virtual DB CRUD) ----------
@api_router.get("/notifications")
async def list_notifications(employee_id: Optional[str] = None, only_due: bool = False):
    """List notifications. If employee_id is provided, returns ones targeting them or 'all'.
    If only_due=true, returns only the ones whose fire_at is now-or-past and not yet marked sent.
    """
    q = {}
    if employee_id:
        q["target"] = {"$in": ["all", employee_id]}
    items = await db.notifications.find(q, {"_id": 0}).sort("fire_at", 1).to_list(2000)
    if only_due:
        now = datetime.now(timezone.utc).isoformat()
        items = [n for n in items if n["fire_at"] <= now and not n.get("sent")]
    return items


@api_router.post("/notifications")
async def create_notification(payload: NotificationCreate):
    """Create a notification record. Auto-creates the 30-min-before-kickoff reminder for a match."""
    match = await db.matches.find_one({"id": payload.match_id})
    if not match:
        raise HTTPException(404, "Match not found")
    n = Notification(**payload.model_dump())
    await db.notifications.insert_one(n.model_dump())
    return n.model_dump()


@api_router.post("/notifications/{notification_id}/sent")
async def mark_notification_sent(notification_id: str):
    res = await db.notifications.update_one({"id": notification_id}, {"$set": {"sent": True}})
    if res.matched_count == 0:
        raise HTTPException(404, "Notification not found")
    return {"ok": True}


@api_router.delete("/notifications/{notification_id}")
async def delete_notification(notification_id: str):
    await db.notifications.delete_one({"id": notification_id})
    return {"ok": True}


@api_router.post("/notifications/auto-schedule", dependencies=[Depends(verify_admin)])
async def auto_schedule_match_reminders(minutes_before: int = 30):
    """For each upcoming match, ensure a single 'kickoff reminder' notification exists
    set to fire `minutes_before` minutes before kickoff."""
    upcoming = await db.matches.find({"status": "upcoming"}, {"_id": 0}).to_list(2000)
    created = 0
    for m in upcoming:
        try:
            ko = datetime.fromisoformat(m["kickoff"].replace("Z", "+00:00"))
        except Exception:
            continue
        fire_at = (ko - timedelta(minutes=minutes_before)).isoformat()
        existing = await db.notifications.find_one({"match_id": m["id"], "title": {"$regex": "Kickoff reminder"}})
        if existing:
            await db.notifications.update_one(
                {"id": existing["id"]},
                {"$set": {"fire_at": fire_at, "sent": False}}
            )
            continue
        n = Notification(
            match_id=m["id"],
            title="Kickoff reminder",
            body=f"{m['team_a']} vs {m['team_b']} starts in {minutes_before} minutes",
            fire_at=fire_at,
            channel="in_app",
            target="all",
        )
        await db.notifications.insert_one(n.model_dump())
        created += 1
    return {"ok": True, "created": created, "total_upcoming": len(upcoming)}


@api_router.get("/notifications/upcoming-for-employee/{employee_id}")
async def upcoming_for_employee(employee_id: str, window_minutes: int = 60):
    """Return notifications relevant to this employee that fire within the next `window_minutes`.
    Only matches the employee has predicted (to avoid noise)."""
    preds = await db.predictions.find({"employee_id": employee_id}, {"_id": 0}).to_list(2000)
    pred_match_ids = {p["match_id"] for p in preds}
    now = datetime.now(timezone.utc)
    until = (now + timedelta(minutes=window_minutes)).isoformat()
    items = await db.notifications.find({
        "match_id": {"$in": list(pred_match_ids)},
        "target": {"$in": ["all", employee_id]},
        "fire_at": {"$lte": until, "$gte": now.isoformat()},
    }, {"_id": 0}).sort("fire_at", 1).to_list(500)
    # Attach match info
    if items:
        mids = list({n["match_id"] for n in items})
        matches = await db.matches.find({"id": {"$in": mids}}, {"_id": 0}).to_list(2000)
        mmap = {m["id"]: m for m in matches}
        for n in items:
            n["match"] = mmap.get(n["match_id"])
    return items


# ---------- Predictions ----------
@api_router.post("/predictions")
async def create_prediction(payload: PredictionCreate):
    emp = await db.employees.find_one({"employee_id": payload.employee_id})
    if not emp:
        raise HTTPException(404, "Employee not registered")
    match = await db.matches.find_one({"id": payload.match_id})
    if not match:
        raise HTTPException(404, "Match not found")
    if is_locked(match):
        raise HTTPException(400, "Predictions are closed (locked 5 minutes before kickoff)")
    if payload.winner not in ("team_a", "team_b", "draw"):
        raise HTTPException(400, "Invalid winner")
    if payload.score_a < 0 or payload.score_b < 0 or payload.score_a > 30 or payload.score_b > 30:
        raise HTTPException(400, "Invalid score")
    # Auto-derive winner from score for consistency
    derived = determine_winner(payload.score_a, payload.score_b)
    winner = derived  # the score determines the winner

    existing = await db.predictions.find_one({
        "employee_id": payload.employee_id,
        "match_id": payload.match_id,
    })
    if existing:
        await db.predictions.update_one(
            {"id": existing["id"]},
            {"$set": {
                "winner": winner,
                "score_a": payload.score_a,
                "score_b": payload.score_b,
            }},
        )
        return clean(await db.predictions.find_one({"id": existing["id"]}))
    pred = Prediction(
        employee_id=payload.employee_id,
        match_id=payload.match_id,
        winner=winner,
        score_a=payload.score_a,
        score_b=payload.score_b,
    )
    await db.predictions.insert_one(pred.model_dump())
    return pred.model_dump()


@api_router.get("/predictions/me")
async def my_predictions(employee_id: str):
    preds = await db.predictions.find({"employee_id": employee_id}, {"_id": 0}).to_list(1000)
    return preds


@api_router.get("/predictions/match/{match_id}")
async def match_predictions(match_id: str):
    """Returns predictions for a match, BUT only if match is finished (anti-cheating)."""
    match = await db.matches.find_one({"id": match_id}, {"_id": 0})
    if not match:
        raise HTTPException(404, "Match not found")
    if match["status"] != "finished":
        return {"locked": True, "predictions": []}
    preds = await db.predictions.find({"match_id": match_id}, {"_id": 0}).to_list(1000)
    # enrich with employee names
    emp_ids = list({p["employee_id"] for p in preds})
    emps = await db.employees.find({"employee_id": {"$in": emp_ids}}, {"_id": 0}).to_list(10000)
    emp_map = {e["employee_id"]: e["name"] for e in emps}
    for p in preds:
        p["employee_name"] = emp_map.get(p["employee_id"], p["employee_id"])
    return {"locked": False, "predictions": preds}


# ---------- Leaderboard ----------
@api_router.get("/leaderboard")
async def leaderboard():
    """
    Anti-cheating: leaderboard only counts points from FINISHED matches.
    Updates only after admin submits result.
    """
    finished = await db.matches.find({"status": "finished"}, {"_id": 0, "id": 1}).to_list(1000)
    finished_ids = [m["id"] for m in finished]
    if not finished_ids:
        return {"entries": [], "finished_matches": 0}
    preds = await db.predictions.find(
        {"match_id": {"$in": finished_ids}}, {"_id": 0}
    ).to_list(100000)
    totals = {}
    for p in preds:
        eid = p["employee_id"]
        totals.setdefault(eid, {"employee_id": eid, "points": 0, "correct_winners": 0, "exact_scores": 0, "total_predictions": 0})
        totals[eid]["points"] += p.get("points", 0)
        totals[eid]["total_predictions"] += 1
        if p.get("points", 0) == 5:
            totals[eid]["exact_scores"] += 1
        elif p.get("points", 0) == 3:
            totals[eid]["correct_winners"] += 1
    # attach names
    emps = await db.employees.find({}, {"_id": 0}).to_list(10000)
    emp_map = {e["employee_id"]: e["name"] for e in emps}
    entries = []
    for eid, t in totals.items():
        t["name"] = emp_map.get(eid, eid)
        entries.append(t)
    entries.sort(key=lambda x: (-x["points"], -x["exact_scores"], x["name"]))
    for i, e in enumerate(entries):
        e["rank"] = i + 1
    return {"entries": entries, "finished_matches": len(finished_ids)}


# ---------- Seed World Cup ----------
# Official FIFA World Cup 2026 final draw (December 5, 2025)
# Format: (team_en, team_ar, flag_emoji)
WC_TEAMS_BY_GROUP = {
    "A": [
        ("Mexico", "المكسيك", "🇲🇽"),
        ("South Africa", "جنوب أفريقيا", "🇿🇦"),
        ("South Korea", "كوريا الجنوبية", "🇰🇷"),
        ("Czech Republic", "التشيك", "🇨🇿"),
    ],
    "B": [
        ("Canada", "كندا", "🇨🇦"),
        ("Bosnia & Herzegovina", "البوسنة والهرسك", "🇧🇦"),
        ("Qatar", "قطر", "🇶🇦"),
        ("Switzerland", "سويسرا", "🇨🇭"),
    ],
    "C": [
        ("Brazil", "البرازيل", "🇧🇷"),
        ("Morocco", "المغرب", "🇲🇦"),
        ("Haiti", "هايتي", "🇭🇹"),
        ("Scotland", "اسكتلندا", "🏴󠁧󠁢󠁳󠁣󠁴󠁿"),
    ],
    "D": [
        ("United States", "الولايات المتحدة", "🇺🇸"),
        ("Paraguay", "باراغواي", "🇵🇾"),
        ("Australia", "أستراليا", "🇦🇺"),
        ("Türkiye", "تركيا", "🇹🇷"),
    ],
    "E": [
        ("Germany", "ألمانيا", "🇩🇪"),
        ("Curaçao", "كوراساو", "🇨🇼"),
        ("Ivory Coast", "ساحل العاج", "🇨🇮"),
        ("Ecuador", "الإكوادور", "🇪🇨"),
    ],
    "F": [
        ("Netherlands", "هولندا", "🇳🇱"),
        ("Japan", "اليابان", "🇯🇵"),
        ("Sweden", "السويد", "🇸🇪"),
        ("Tunisia", "تونس", "🇹🇳"),
    ],
    "G": [
        ("Belgium", "بلجيكا", "🇧🇪"),
        ("Egypt", "مصر", "🇪🇬"),
        ("Iran", "إيران", "🇮🇷"),
        ("New Zealand", "نيوزيلندا", "🇳🇿"),
    ],
    "H": [
        ("Spain", "إسبانيا", "🇪🇸"),
        ("Cape Verde", "الرأس الأخضر", "🇨🇻"),
        ("Saudi Arabia", "السعودية", "🇸🇦"),
        ("Uruguay", "أوروغواي", "🇺🇾"),
    ],
    "I": [
        ("France", "فرنسا", "🇫🇷"),
        ("Senegal", "السنغال", "🇸🇳"),
        ("Iraq", "العراق", "🇮🇶"),
        ("Norway", "النرويج", "🇳🇴"),
    ],
    "J": [
        ("Argentina", "الأرجنتين", "🇦🇷"),
        ("Algeria", "الجزائر", "🇩🇿"),
        ("Austria", "النمسا", "🇦🇹"),
        ("Jordan", "الأردن", "🇯🇴"),
    ],
    "K": [
        ("Portugal", "البرتغال", "🇵🇹"),
        ("DR Congo", "الكونغو الديمقراطية", "🇨🇩"),
        ("Uzbekistan", "أوزبكستان", "🇺🇿"),
        ("Colombia", "كولومبيا", "🇨🇴"),
    ],
    "L": [
        ("England", "إنجلترا", "🏴󠁧󠁢󠁥󠁮󠁧󠁿"),
        ("Croatia", "كرواتيا", "🇭🇷"),
        ("Ghana", "غانا", "🇬🇭"),
        ("Panama", "بنما", "🇵🇦"),
    ],
}


def _generate_group_stage_matches():
    """
    Generate 72 group-stage matches across June 11–27, 2026.
    Standard round-robin pairing per group: (1v2, 3v4), (1v3, 2v4), (1v4, 2v3).
    Matchday 1: Jun 11–16 (24 matches over 6 days, ~4/day)
    Matchday 2: Jun 17–22
    Matchday 3: Jun 23–27 (last 2 days have simultaneous kickoffs)
    Opening match: Mexico vs South Africa, June 11 at 20:00 UTC (Estadio Azteca).
    """
    pairings = [(0, 1), (2, 3), (0, 2), (1, 3), (0, 3), (1, 2)]
    groups = list(WC_TEAMS_BY_GROUP.keys())  # A..L
    # 12 groups × 6 matchdays-pairings -> 72 entries, but pairings are 2 per matchday => need careful schedule
    # Easier: flatten as (group, p1_idx, p2_idx, matchday) — matchday 1: pairings 0&1; matchday 2: 2&3; matchday 3: 4&5
    venues = [
        "Estadio Azteca, Mexico City", "BMO Field, Toronto", "MetLife Stadium, NJ",
        "SoFi Stadium, Los Angeles", "AT&T Stadium, Dallas", "Lincoln Financial Field, Philadelphia",
        "Hard Rock Stadium, Miami", "Mercedes-Benz Stadium, Atlanta", "Lumen Field, Seattle",
        "Levi's Stadium, San Francisco", "Arrowhead Stadium, Kansas City", "Gillette Stadium, Boston",
        "NRG Stadium, Houston", "BC Place, Vancouver", "Estadio BBVA, Monterrey",
        "Estadio Akron, Guadalajara",
    ]
    base_day = datetime(2026, 6, 11, tzinfo=timezone.utc)
    matches = []
    # For each matchday (1,2,3), iterate through groups and emit 2 matches per group
    # Matchday 1: days 0-5; Matchday 2: days 6-11; Matchday 3: days 12-16
    matchday_day_offsets = {1: list(range(0, 6)), 2: list(range(6, 12)), 3: list(range(12, 17))}
    kickoff_slots = ["17:00", "20:00", "22:30", "01:00"]  # UTC time slots
    counter = 0
    for md in (1, 2, 3):
        md_pairings = pairings[(md - 1) * 2:(md - 1) * 2 + 2]
        day_offsets = matchday_day_offsets[md]
        for gi, g in enumerate(groups):
            teams = WC_TEAMS_BY_GROUP[g]
            for pi, (a_idx, b_idx) in enumerate(md_pairings):
                # Distribute across days/slots
                slot_idx = counter % len(kickoff_slots)
                day_idx = (counter // 4) % len(day_offsets)
                date = base_day + timedelta(days=day_offsets[day_idx])
                hh, mm = kickoff_slots[slot_idx].split(":")
                ko = date.replace(hour=int(hh), minute=int(mm))
                a = teams[a_idx]
                b = teams[b_idx]
                # Mexico opens at Azteca on June 11 20:00 UTC — special-case
                if md == 1 and g == "A" and a[0] == "Mexico" and b[0] == "South Africa":
                    ko = datetime(2026, 6, 11, 20, 0, tzinfo=timezone.utc)
                venue = venues[(gi + pi + md) % len(venues)]
                matches.append({
                    "group": g,
                    "team_a_en": a[0], "team_a_ar": a[1], "flag_a": a[2],
                    "team_b_en": b[0], "team_b_ar": b[1], "flag_b": b[2],
                    "kickoff": ko.isoformat().replace("+00:00", "Z"),
                    "venue": venue,
                })
                counter += 1
    # Sort by kickoff
    matches.sort(key=lambda m: m["kickoff"])
    return matches


@api_router.post("/admin/seed", dependencies=[Depends(verify_admin)])
async def seed():
    count = await db.matches.count_documents({})
    if count > 0:
        return {"ok": True, "skipped": True, "existing": count}
    docs = []
    for spec in _generate_group_stage_matches():
        m = Match(
            team_a=spec["team_a_en"], team_b=spec["team_b_en"],
            team_a_ar=spec["team_a_ar"], team_b_ar=spec["team_b_ar"],
            flag_a=spec["flag_a"], flag_b=spec["flag_b"],
            group=spec["group"], stage="Group Stage",
            kickoff=spec["kickoff"], venue=spec["venue"],
        )
        docs.append(m.model_dump())
    await db.matches.insert_many(docs)
    return {"ok": True, "inserted": len(docs)}


@api_router.post("/admin/reset", dependencies=[Depends(verify_admin)])
async def reset():
    await db.matches.delete_many({})
    await db.predictions.delete_many({})
    return {"ok": True}


@api_router.get("/admin/check")
async def admin_check(x_admin_password: str = Header(None)):
    return {"ok": x_admin_password == ADMIN_PASSWORD}


@api_router.post("/admin/matches/{match_id}/stream", dependencies=[Depends(verify_admin)])
async def set_stream_url(match_id: str, stream_url: str = ""):
    """Set or clear an embeddable stream URL (e.g., YouTube embed) for a match."""
    res = await db.matches.update_one(
        {"id": match_id}, {"$set": {"stream_url": stream_url.strip() or None}}
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Match not found")
    return {"ok": True}


def _fmt_en_datetime(iso: str) -> str:
    """Format ISO timestamp as full English: 'Wednesday, 17 June 2026, 08:30:42 PM (UTC)'."""
    try:
        d = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return d.strftime("%A, %d %B %Y, %I:%M:%S %p (UTC)")
    except Exception:
        return iso


@api_router.get("/admin/matches/{match_id}/export.xlsx", dependencies=[Depends(verify_admin)])
async def export_match_predictions_xlsx(match_id: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    import io

    match = await db.matches.find_one({"id": match_id}, {"_id": 0})
    if not match:
        raise HTTPException(404, "Match not found")
    preds = await db.predictions.find({"match_id": match_id}, {"_id": 0}).sort("created_at", 1).to_list(20000)
    emp_ids = list({p["employee_id"] for p in preds})
    emps = await db.employees.find({"employee_id": {"$in": emp_ids}}, {"_id": 0}).to_list(20000)
    emp_map = {e["employee_id"]: e["name"] for e in emps}

    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A2B6D")

    # Match info header
    ws["A1"] = f"Match: {match['team_a']} vs {match['team_b']}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")
    ws["A2"] = f"Group {match['group']} · {_fmt_en_datetime(match['kickoff'])}"
    ws["A2"].font = Font(italic=True, color="555555")
    ws.merge_cells("A2:F2")
    if match.get("status") == "finished":
        ws["A3"] = f"Final score: {match['result_a']} - {match['result_b']}"
        ws["A3"].font = Font(bold=True, color="007A3D")
        ws.merge_cells("A3:F3")

    # Column headers
    headers = ["#", "Employee Name", "Employee ID", "Prediction (A-B)", "Submitted At (English)", "Points"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for i, p in enumerate(preds, 1):
        ws.cell(row=5 + i, column=1, value=i)
        ws.cell(row=5 + i, column=2, value=emp_map.get(p["employee_id"], p["employee_id"]))
        ws.cell(row=5 + i, column=3, value=p["employee_id"])
        ws.cell(row=5 + i, column=4, value=f"{p['score_a']} - {p['score_b']}").alignment = Alignment(horizontal="center")
        ws.cell(row=5 + i, column=5, value=_fmt_en_datetime(p["created_at"]))
        ws.cell(row=5 + i, column=6, value=p.get("points", 0)).alignment = Alignment(horizontal="center")

    widths = [6, 28, 16, 18, 42, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = f"predictions_{match['team_a']}_vs_{match['team_b']}.xlsx".replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )


@api_router.get("/admin/predictions/export.xlsx", dependencies=[Depends(verify_admin)])
async def export_all_predictions_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    import io

    matches = await db.matches.find({}, {"_id": 0}).to_list(2000)
    match_map = {m["id"]: m for m in matches}
    preds = await db.predictions.find({}, {"_id": 0}).sort("created_at", 1).to_list(200000)
    emp_ids = list({p["employee_id"] for p in preds})
    emps = await db.employees.find({"employee_id": {"$in": emp_ids}}, {"_id": 0}).to_list(20000)
    emp_map = {e["employee_id"]: e["name"] for e in emps}

    wb = Workbook()
    ws = wb.active
    ws.title = "All Predictions"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A2B6D")
    title = ws.cell(row=1, column=1, value="NCC World Cup 2026 — All Predictions Export")
    title.font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")
    ws.cell(row=2, column=1, value=f"Generated: {_fmt_en_datetime(datetime.now(timezone.utc).isoformat())}").font = Font(italic=True, color="555555")
    ws.merge_cells("A2:I2")

    headers = [
        "#", "Employee Name", "Employee ID", "Group", "Match",
        "Kickoff (English)", "Prediction", "Submitted At (English)", "Points",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for i, p in enumerate(preds, 1):
        m = match_map.get(p["match_id"], {})
        ws.cell(row=4 + i, column=1, value=i)
        ws.cell(row=4 + i, column=2, value=emp_map.get(p["employee_id"], p["employee_id"]))
        ws.cell(row=4 + i, column=3, value=p["employee_id"])
        ws.cell(row=4 + i, column=4, value=m.get("group", ""))
        ws.cell(row=4 + i, column=5, value=f"{m.get('team_a', '?')} vs {m.get('team_b', '?')}")
        ws.cell(row=4 + i, column=6, value=_fmt_en_datetime(m.get("kickoff", "")))
        ws.cell(row=4 + i, column=7, value=f"{p['score_a']} - {p['score_b']}")
        ws.cell(row=4 + i, column=8, value=_fmt_en_datetime(p["created_at"]))
        ws.cell(row=4 + i, column=9, value=p.get("points", 0))

    widths = [6, 26, 14, 8, 32, 42, 12, 42, 8]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ncc_all_predictions.xlsx"'},
    )


@api_router.get("/")
async def root():
    return {"message": "NCC World Cup Contest API"}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


@app.on_event("startup")
async def auto_seed():
    try:
        # Clean up legacy employees missing username/password_hash (pre-auth migration)
        legacy = await db.employees.count_documents({"$or": [
            {"username": {"$exists": False}},
            {"password_hash": {"$exists": False}},
        ]})
        if legacy:
            await db.employees.delete_many({"$or": [
                {"username": {"$exists": False}},
                {"password_hash": {"$exists": False}},
            ]})
            # Also drop predictions for non-existent employees (orphan cleanup)
            valid_ids = [e["employee_id"] async for e in db.employees.find({}, {"employee_id": 1, "_id": 0})]
            await db.predictions.delete_many({"employee_id": {"$nin": valid_ids}})
            logger.info(f"Migration: removed {legacy} legacy employees + orphan predictions")

        # Unique indexes
        await db.employees.create_index("username", unique=True)
        await db.employees.create_index("employee_id", unique=True)

        count = await db.matches.count_documents({})
        if count == 0:
            docs = []
            for spec in _generate_group_stage_matches():
                m = Match(
                    team_a=spec["team_a_en"], team_b=spec["team_b_en"],
                    team_a_ar=spec["team_a_ar"], team_b_ar=spec["team_b_ar"],
                    flag_a=spec["flag_a"], flag_b=spec["flag_b"],
                    group=spec["group"], stage="Group Stage",
                    kickoff=spec["kickoff"], venue=spec["venue"],
                )
                docs.append(m.model_dump())
            await db.matches.insert_many(docs)
            logger.info(f"Auto-seeded {len(docs)} World Cup matches")
    except Exception as e:
        logger.error(f"Seed error: {e}")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
