"""Backend tests for NCC World Cup Contest - Iteration 3.
Covers:
- 72-match seed across 12 groups (A..L)
- Saudi Arabia in Group H (with Spain, Cape Verde, Uruguay)
- Stream URL endpoint
- XLSX exports (per-match and all predictions)
- Regression: no break in registration, prediction submit, leaderboard, latest winners
"""
import io
import os
import uuid
from collections import defaultdict
from pathlib import Path

import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL")
if not BASE_URL:
    env_file = Path("/app/frontend/.env")
    for line in env_file.read_text().splitlines():
        if line.startswith("REACT_APP_BACKEND_URL="):
            BASE_URL = line.split("=", 1)[1].strip()
            break
BASE_URL = BASE_URL.rstrip("/")
API = f"{BASE_URL}/api"
ADMIN_PASSWORD = "ncc-admin-2026"

RUN = uuid.uuid4().hex[:6]
EID = f"TEST_I3_E_{RUN}"


@pytest.fixture(scope="module")
def session():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    return s


@pytest.fixture(scope="module")
def admin_headers():
    return {"X-Admin-Password": ADMIN_PASSWORD, "Content-Type": "application/json"}


@pytest.fixture(scope="module")
def all_matches(session):
    r = session.get(f"{API}/matches")
    assert r.status_code == 200, r.text
    return r.json()


# ---------- Seed: 72 matches across 12 groups ----------
class TestSeed72Groups:
    def test_total_match_count_72(self, all_matches):
        assert len(all_matches) == 72, f"Expected 72 matches, got {len(all_matches)}"

    def test_twelve_groups_a_to_l(self, all_matches):
        groups = sorted({m["group"] for m in all_matches})
        expected = list("ABCDEFGHIJKL")
        assert groups == expected, f"Groups mismatch: {groups}"

    def test_each_group_has_six_matches(self, all_matches):
        counts = defaultdict(int)
        for m in all_matches:
            counts[m["group"]] += 1
        for g in "ABCDEFGHIJKL":
            assert counts[g] == 6, f"Group {g} has {counts[g]} matches, expected 6"

    def test_each_match_has_venue_and_kickoff(self, all_matches):
        for m in all_matches:
            assert m.get("kickoff"), f"Match {m['id']} missing kickoff"
            assert m.get("venue"), f"Match {m['id']} missing venue"
            assert m.get("team_a_ar") and m.get("team_b_ar"), "Arabic names missing"


# ---------- Saudi Arabia in Group H ----------
class TestSaudiArabiaGroupH:
    def test_saudi_arabia_three_matches_in_group_h(self, all_matches):
        sa_matches = [
            m for m in all_matches
            if m["group"] == "H" and (
                m["team_a"] == "Saudi Arabia" or m["team_b"] == "Saudi Arabia"
            )
        ]
        assert len(sa_matches) == 3, f"Expected 3 Saudi matches in H, got {len(sa_matches)}"
        for m in sa_matches:
            assert "السعودية" in (m["team_a_ar"], m["team_b_ar"]), \
                f"Arabic name missing on {m}"
            assert m["venue"], "Venue must be populated"

    def test_group_h_includes_spain_cape_verde_saudi_uruguay(self, all_matches):
        h_matches = [m for m in all_matches if m["group"] == "H"]
        teams = set()
        for m in h_matches:
            teams.add(m["team_a"])
            teams.add(m["team_b"])
        expected = {"Spain", "Cape Verde", "Saudi Arabia", "Uruguay"}
        assert expected.issubset(teams), f"Group H teams = {teams}, missing {expected - teams}"


# ---------- Stream URL ----------
class TestStreamUrl:
    def test_set_stream_url_admin(self, session, admin_headers, all_matches):
        m = all_matches[0]
        yt = "https://www.youtube.com/watch?v=dQw4w9WgXcQ"
        r = session.post(
            f"{API}/admin/matches/{m['id']}/stream",
            headers=admin_headers,
            params={"stream_url": yt},
        )
        assert r.status_code == 200, r.text
        # GET match returns it
        g = session.get(f"{API}/matches/{m['id']}").json()
        assert g["stream_url"] == yt
        # cleanup
        session.post(
            f"{API}/admin/matches/{m['id']}/stream",
            headers=admin_headers,
            params={"stream_url": ""},
        )
        g2 = session.get(f"{API}/matches/{m['id']}").json()
        assert g2.get("stream_url") in (None, "")

    def test_set_stream_url_requires_admin(self, session, all_matches):
        m = all_matches[0]
        r = session.post(
            f"{API}/admin/matches/{m['id']}/stream",
            params={"stream_url": "https://example.com/x"},
        )
        assert r.status_code == 401

    def test_set_stream_url_404_unknown_match(self, session, admin_headers):
        r = session.post(
            f"{API}/admin/matches/does-not-exist/stream",
            headers=admin_headers,
            params={"stream_url": "https://x.com/"},
        )
        assert r.status_code == 404


# ---------- XLSX Exports ----------
class TestXlsxExports:
    XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def test_match_export_requires_admin(self, session, all_matches):
        m = all_matches[0]
        r = session.get(f"{API}/admin/matches/{m['id']}/export.xlsx")
        assert r.status_code == 401

    def test_match_export_returns_xlsx(self, session, admin_headers, all_matches):
        m = all_matches[0]
        r = session.get(
            f"{API}/admin/matches/{m['id']}/export.xlsx",
            headers=admin_headers,
        )
        assert r.status_code == 200, r.text
        assert r.headers["content-type"].startswith(self.XLSX_CT)
        # Must be a real ZIP (xlsx) file
        assert r.content[:4] == b"PK\x03\x04", "Not a valid ZIP/XLSX magic"
        assert len(r.content) > 1000
        # Open and verify content
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        assert ws["A1"].value and "Match:" in ws["A1"].value
        # English date format in row 2
        assert ws["A2"].value and "Group" in ws["A2"].value

    def test_all_predictions_export_requires_admin(self, session):
        r = session.get(f"{API}/admin/predictions/export.xlsx")
        assert r.status_code == 401

    def test_all_predictions_export_returns_xlsx(self, session, admin_headers):
        r = session.get(f"{API}/admin/predictions/export.xlsx", headers=admin_headers)
        assert r.status_code == 200, r.text
        assert r.headers["content-type"].startswith(self.XLSX_CT)
        assert r.content[:4] == b"PK\x03\x04"
        # Verify it opens and has the expected headers
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        title = ws["A1"].value or ""
        assert "All Predictions" in title or "NCC" in title
        # Header row at row 4
        headers_row = [ws.cell(row=4, column=c).value for c in range(1, 10)]
        assert "Employee Name" in headers_row
        assert "Employee ID" in headers_row
        assert "Group" in headers_row
        assert "Points" in headers_row


# ---------- Regression: registration / prediction / leaderboard / latest winners ----------
class TestRegression:
    def test_register_and_login(self, session):
        r = session.post(
            f"{API}/auth/register",
            json={"employee_id": EID, "name": "Iter3 Tester"},
        )
        assert r.status_code == 200
        assert r.json()["employee_id"] == EID
        r2 = session.post(f"{API}/auth/login", json={"employee_id": EID})
        assert r2.status_code == 200

    def test_submit_prediction_on_upcoming_match(self, session, all_matches):
        # find a match safely far in the future
        from datetime import datetime, timezone
        upcoming = [
            m for m in all_matches
            if m["status"] == "upcoming"
            and datetime.fromisoformat(m["kickoff"].replace("Z", "+00:00")) > datetime.now(timezone.utc)
        ]
        assert upcoming, "Need upcoming matches"
        m = upcoming[0]
        # ensure employee exists
        session.post(f"{API}/auth/register", json={"employee_id": EID, "name": "Iter3 Tester"})
        r = session.post(
            f"{API}/predictions",
            json={
                "employee_id": EID,
                "match_id": m["id"],
                "winner": "team_a",
                "score_a": 1,
                "score_b": 0,
            },
        )
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["score_a"] == 1 and d["score_b"] == 0
        assert d["winner"] == "team_a"

    def test_leaderboard_endpoint_works(self, session):
        r = session.get(f"{API}/leaderboard")
        assert r.status_code == 200
        body = r.json()
        assert "entries" in body and "finished_matches" in body

    def test_latest_winners_endpoint_works(self, session):
        r = session.get(f"{API}/winners/latest")
        assert r.status_code == 200
        body = r.json()
        assert "match" in body and "winners" in body
